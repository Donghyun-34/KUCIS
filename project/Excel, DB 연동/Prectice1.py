import openpyxl
import pymysql


class test:
    def __init__(self, num, name):
        self.num = num
        self.name = name


def get_connection():
    conn = pymysql.connect(host='localhost',
                           user='root',
                           password='ehdgus123',
                           db='python',
                           charset='utf8')
    return conn


def insert(object):
    conn = get_connection()
    try:
        with conn.cursor() as curs:
            sql = 'insert into test values(%s, %s)'
            curs.execute(sql, (object.num, object.name))
        conn.commit()
    finally:
        conn.close()


def select_all():
    conn = get_connection()
    try:
        with conn.cursor() as curs:
            sql = "select * from test"
            curs.execute(sql)
            rs = curs.fetchall()
            for row in rs:
                print(row)
    finally:
        conn.close()


def delete(num):
    conn = get_connection()
    try:
        with conn.cursor() as curs:
            sql = 'delete from test where num=%s'
            curs.execute(sql, num)
        conn.commit()
    finally:
        conn.close()


def delete_all():
    conn = get_connection()
    try:
        with conn.cursor() as curs:
            sql = 'delete from test'
            curs.execute(sql)
        conn.commit()
    finally:
        conn.close()


def update(object):
    conn = get_connection()
    try:
        with conn.cursor() as curs:
            sql = 'update test set name=%s where num=%d'
            curs.execute(sql, object.name, object.num)
        conn.commit()
    finally:
        conn.close()

def insert_excel_to_db():
    conn = get_connection()
    try:
        with conn.cursor() as curs:
            sql = 'insert into test values (%s, %s)'

            wb = openpyxl.load_workbook('excel_sql.xlsx', data_only=True)
            ws = wb['Sheet1']

            for row in ws.iter_rows(max_col= 3, max_row=12):
                try:
                    curs.execute(sql, (row[0].value, row[1].value))
                except:
                    continue
            conn.commit()

    finally:
        conn.close()
        wb.close()


def search_from_excel_and_update_to_db():
    wb = openpyxl.load_workbook('excel_sql.xlsx', data_only=True)
    ws = wb['Sheet1']

    for row in ws.iter_rows(max_col=3, max_row=12):
        try:
            if row[1].value == 'Kim Donghyeon':
                tmp = test(row[0].value, 'Kim Dong')
                update(tmp)
                break
        except:
            continue

    wb.close()


search_from_excel_and_update_to_db()
select_all()
